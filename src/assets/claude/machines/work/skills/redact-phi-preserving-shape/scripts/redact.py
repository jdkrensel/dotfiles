"""Apply a column plan to a delimited or Excel file, preserving its shape exactly.

Usage:
    uv run --with openpyxl python redact.py --plan plan.json --in FILE --out FILE

The division of labour is deliberate. Deciding *which* columns hold PHI needs
judgement about clinical data, so a human-or-model writes that decision down in
a plan file. Applying the decision to every cell consistently is mechanical, and
mechanical work belongs in code -- a script does not get bored on row 4,000 and
silently leave a name unredacted.

Everything not named in the plan is copied through byte-for-byte. That default
matters: clinical codes, payors, procedure text and the file's own structure are
what make a redacted file useful for testing, and quietly mangling them would
defeat the purpose.
"""

from __future__ import annotations

import argparse
import csv
import datetime
import io
import json
import os
import secrets
import shutil
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

import shapes  # noqa: E402
from shapes import Persona, _rand, fill_mask, mask_of, normalize_key, parse_date, shift_date  # noqa: E402

SALT_ENV = "REDACT_SALT"


class Redactor:
    """Holds the identity mappings so the same input always yields the same fake.

    Consistency across a file is not a nicety. Duplicate-patient handling,
    join-on-MRN logic, and repeat-visit ordering are all things a client file
    exercises, and they only survive redaction if a given real value maps to one
    fake value everywhere it appears.
    """

    def __init__(self, plan: dict, salt: str) -> None:
        self.plan = plan
        self.salt = salt
        self.columns: dict[str, dict] = {}
        for entry in plan.get("columns", []):
            for name in [entry["name"]] + list(entry.get("aliases", [])):
                self.columns[name.strip().lower()] = entry
        self.identity_key_cols = [c.strip().lower() for c in plan.get("identity_key", [])]
        self._personas: dict[str, Persona] = {}
        self._values: dict[tuple[str, str], str] = {}
        self._offsets: dict[str, int] = {}
        self.stats: dict[str, int] = {}
        # Every name word appearing anywhere in the source. Generated names steer
        # clear of all of them, so no invented name coincides with a real one even
        # from a different patient. Without this, a common-name pool collides by
        # chance and the output contains real names -- harmless in principle, but
        # indistinguishable from a genuine leak to anyone auditing the file.
        self.forbidden_words: set[str] = set()

    # -- identity ---------------------------------------------------------

    def _row_identity(self, row: dict[str, str]) -> str:
        """The subject key for a row: what makes two rows the same person.

        Prefer a stable identifier (MRN) over a name, because names are spelled
        inconsistently in source systems while MRNs are not.
        """
        parts = []
        for col in self.identity_key_cols:
            val = (row.get(col) or "").strip()
            if val:
                parts.append(normalize_key(val))
        if parts:
            return "|".join(parts)
        # No identity column produced a value -- fall back to whole-row identity so
        # each such row gets its own persona rather than all sharing one.
        return "row:" + normalize_key("|".join(f"{k}={v}" for k, v in sorted(row.items())))

    def collect_name_words(self, rows, name_cols) -> None:
        """Pre-scan name columns so generation can avoid every real word."""
        for row in rows:
            for col in name_cols:
                val = row.get(col) or ""
                for part in str(val).replace(",", " ").split():
                    bare = part.rstrip(".").upper()
                    if len(bare) >= 3:
                        self.forbidden_words.add(bare)

    def persona(self, subject: str, first_len: int = 7, last_len: int = 8) -> Persona:
        if subject not in self._personas:
            self._personas[subject] = Persona(
                _rand(self.salt + ":persona", subject), first_len, last_len,
                forbidden=self.forbidden_words)
        return self._personas[subject]

    def date_offset(self, subject: str) -> int:
        """One shift per subject, so their dates move together.

        Shifting every date for a person by the same amount keeps the intervals
        that clinical logic reads -- age at surgery, days between visits, the
        order of encounters -- while breaking the link to the real calendar.
        """
        if subject not in self._offsets:
            lo, hi = self.plan.get("date_offset_range", [-540, 540])
            rng = _rand(self.salt + ":offset", subject)
            span = hi - lo
            off = lo + rng.below(span + 1)
            # A zero shift would leave the real date in place.
            if off == 0:
                off = 31
            self._offsets[subject] = off
        return self._offsets[subject]

    # -- cell dispatch ----------------------------------------------------

    def redact_cell(self, col: str, value: str, subject: str) -> str:
        entry = self.columns.get(col.strip().lower())
        if entry is None:
            return value  # not PHI: preserve exactly
        if value is None:
            return value
        raw = str(value)
        if raw.strip() == "":
            return raw  # empty stays empty; blankness is part of the shape

        kind = entry.get("type", "id")
        if kind == "keep":
            # An explicit "I looked at this column and it is not PHI". Recording the
            # decision costs nothing and is what lets verify.py tell a reviewed
            # column apart from one nobody ever considered.
            return raw
        cache_key = (col.strip().lower() + "|" + kind, raw)
        # Names and dates depend on the subject, so they are cached per subject;
        # opaque ids are globally consistent by value.
        if kind in {"name", "date", "age"}:
            cache_key = (cache_key[0] + "|" + subject, raw)
        if cache_key in self._values:
            return self._values[cache_key]

        rng = _rand(self.salt + ":" + kind + ":" + col.strip().lower(), raw, subject if kind in {"name", "date", "age"} else "")
        result = self._generate(kind, raw, rng, subject, entry)
        self._values[cache_key] = result
        self.stats[kind] = self.stats.get(kind, 0) + 1
        return result

    def _name_part(self, persona: Persona, part: str, value: str) -> str:
        """Fake one component of a name split across columns.

        The persona is shared with any whole-name column, so a file carrying both
        "PAT_NAME" and "First Name"/"Last Name" stays internally consistent.
        """
        bare = value.rstrip(".")
        dot = "." if value.endswith(".") else ""
        role = "last" if part == "last" else "given"
        slot = 1 if part == "middle" else 0
        if len(bare) == 1:
            # A bare initial. It has to change -- an unchanged initial is an
            # unredacted identifier, however small -- while staying one letter and
            # avoiding letters that read as suffixes.
            src = persona.last if role == "last" else (persona.middle if slot else persona.first)
            letter = next((c for c in src if c.upper() != bare.upper()
                           and c.upper() not in shapes._SINGLE_LETTER_SUFFIXES), None)
            if letter is None:
                letter = next(c for c in "BCDEFGHJKLMNPRSTW" if c.upper() != bare.upper())
            return shapes._case_like(letter, value) + dot
        return persona.word(slot, bare, role) + dot

    def _generate(self, kind: str, raw: str, rng, subject: str, entry: dict) -> str:
        stripped = raw.strip()
        lead = raw[: len(raw) - len(raw.lstrip())]
        trail = raw[len(raw.rstrip()):]

        def wrap(s: str) -> str:
            return f"{lead}{s}{trail}"

        if kind == "name":
            persona = self.persona(subject, *_name_lengths(stripped))
            part = entry.get("name_part")
            if part in {"first", "last", "middle"}:
                # A split name column holds one word, so it is sized from this cell
                # rather than from a whole-name guess -- otherwise a 4-letter
                # "TEST" becomes an 8-letter surname and the mask breaks.
                return wrap(self._name_part(persona, part, stripped))
            order = entry.get("name_order")
            last_first = None if order is None else (order == "last_first")
            return wrap(shapes.render_name(stripped, persona, last_first))

        if kind == "date":
            parsed = parse_date(stripped)
            if parsed is None:
                return raw  # unrecognised: leave rather than corrupt
            return wrap(shift_date(parsed, self.date_offset(subject)))

        if kind == "ssn":
            return wrap(shapes.fake_ssn(stripped, rng))
        if kind == "phone":
            return wrap(shapes.fake_phone(stripped, rng))
        if kind == "email":
            return wrap(shapes.fake_email(stripped, rng, self.persona(subject)))
        if kind == "zip":
            return wrap(shapes.fake_zip(stripped, rng))
        if kind == "age":
            return wrap(shapes.fake_age(stripped, rng))
        if kind == "address":
            return wrap(_fake_street(stripped, rng))
        if kind == "city":
            return wrap(shapes._case_like(rng.pick(_CITIES), stripped))
        if kind == "url":
            return wrap(_fake_url(stripped, rng))
        if kind == "ip":
            return wrap(_fake_ip(stripped, rng))
        # "id" and anything unknown: mask-fill, which is shape-exact and safe.
        return wrap(fill_mask(mask_of(stripped), rng, stripped))


def _name_lengths(value: str) -> tuple[int, int]:
    """Approximate the given/family word lengths so the fake matches density.

    Only used to seed the persona's default words; per-cell generation always
    resizes to the actual value, so an imperfect guess here costs nothing.
    """
    words = [w for w in shapes._NAME_SPLIT.split(value)
             if w and shapes._NAME_SPLIT.fullmatch(w) and len(w.rstrip(".")) > 1]
    if not words:
        return 7, 8
    if len(words) == 1:
        return len(words[0]), len(words[0])
    if "," in value:
        return len(words[1]), len(words[0])
    return len(words[0]), len(words[-1])


_CITIES = ["FAIRVIEW", "RIVERTON", "OAKDALE", "SPRINGDALE", "MAPLETON", "CLEARWATER",
           "BRIDGEPORT", "GREENFIELD", "LAKESIDE", "MILLBROOK", "NORTHGATE", "STONEHAVEN"]
_STREETS = ["OAK", "ELM", "ASH", "BAY", "FIR", "MAPLE", "CEDAR", "BIRCH", "ASPEN", "OLIVE",
            "WALNUT", "LAUREL", "WILLOW", "POPLAR", "JUNIPER", "HICKORY", "DOGWOOD",
            "HAWTHORN", "SYCAMORE", "CHESTNUT", "MAGNOLIA", "BASSWOOD"]
_STREET_TYPES = ["ST", "AVE", "RD", "DR", "LN", "CT", "BLVD", "WAY", "PL", "TER"]


# Address words that describe a location's structure rather than name it. Keeping
# them verbatim is what makes the fake address parse the same way: an address
# standardiser keys on "APT", "STE" and the street-type suffix, so replacing them
# would change which code path the row exercises.
_ADDRESS_STRUCTURAL = set(_STREET_TYPES) | {
    "STREET", "AVENUE", "ROAD", "DRIVE", "LANE", "COURT", "BOULEVARD", "PLACE", "TERRACE",
    "CIRCLE", "PARKWAY", "HIGHWAY", "TRAIL", "APT", "APARTMENT", "UNIT", "STE", "SUITE",
    "BLDG", "BUILDING", "FLOOR", "FL", "RM", "ROOM", "PO", "BOX", "N", "S", "E", "W",
    "NE", "NW", "SE", "SW", "NORTH", "SOUTH", "EAST", "WEST",
    # Rural and numbered-route designators: "CR 431", "HWY 51", "FM 1960". These are
    # road classes, not names, and a road number is handled as a number below.
    "CR", "CO", "COUNTY", "RD", "HWY", "FM", "RR", "RT", "RTE", "ROUTE", "US", "SR",
}
# Deliberately NOT structural: "STATE" and similar words that read as road classes but
# are far more often the actual street name ("N STATE ST"). Keeping such a word would
# leave a real street name in the output, so they fall through to replacement.


def _fake_street(source: str, rng) -> str:
    """Invent a street address whose token template mirrors the original."""
    out = []
    for tok in source.split():
        bare = tok.upper().rstrip(".,")
        if tok.isdigit():
            # House and route numbers keep their digit count, so a 5-digit house
            # number does not become a 3-digit one.
            lo = 10 ** (len(tok) - 1) if len(tok) > 1 else 1
            out.append(f"{rng.between(lo, 10 ** len(tok) - 1)}")
        elif bare in _ADDRESS_STRUCTURAL:
            out.append(tok)
        elif any(c.isalpha() for c in tok) and any(c.isdigit() for c in tok):
            out.append(fill_mask(mask_of(tok), rng, tok))  # e.g. "12B"
        elif any(c.isalpha() for c in tok):
            word = rng.pick([s for s in _STREETS if len(s) == len(bare)] or _STREETS)
            out.append(shapes._case_like(word, tok))
        else:
            out.append(fill_mask(mask_of(tok), rng, tok))
    return " ".join(out)


def _fake_url(source: str, rng) -> str:
    scheme, sep, rest = source.partition("://")
    if not sep:
        scheme, sep, rest = "https", "://", source
    path = "/" + fill_mask(mask_of(rest.partition("/")[2] or "abc123"), rng, "").lstrip("/")
    return f"{scheme}{sep}{shapes.RESERVED_DOMAINS[0]}{path}"


def _fake_ip(source: str, rng) -> str:
    """Use the 203.0.113.0/24 documentation range (RFC 5737)."""
    if ":" in source:
        return "2001:db8::" + format(rng.below(65536), "x")
    return f"203.0.113.{rng.between(1, 254)}"


# ---------------------------------------------------------------------------
# Delimited files
# ---------------------------------------------------------------------------

def sniff_text(path: Path, header_row: int = 0) -> dict:
    """Detect encoding, BOM, newline style and delimiter without guessing blind.

    The delimiter is counted on the *header* line rather than the first physical
    line. Files with banner rows above the headers are common, and a banner is
    prose -- sniffing it picks whatever punctuation the sentence happened to use,
    every row then parses as one wide field, no header matches the plan, and the
    file copies through completely unredacted while every shape check still passes.
    """
    raw = path.read_bytes()
    bom = raw.startswith(b"\xef\xbb\xbf")
    body = raw[3:] if bom else raw
    encoding = "utf-8"
    for enc in ("utf-8", "cp1252", "latin-1"):
        try:
            text = body.decode(enc)
            encoding = enc
            break
        except UnicodeDecodeError:
            continue
    else:
        text = body.decode("latin-1")
        encoding = "latin-1"

    newline = "\r\n" if "\r\n" in text else ("\r" if "\r" in text and "\n" not in text else "\n")
    lines = text.split("\n")
    probe = lines[header_row] if 0 <= header_row < len(lines) else (lines[0] if lines else "")
    counts = {d: probe.count(d) for d in ["|", "\t", ",", ";"]}
    delimiter = max(counts, key=lambda d: counts[d]) if max(counts.values()) > 0 else ","
    final_newline = text.endswith(("\n", "\r"))
    # A fully-quoted export must come back fully quoted. Rewriting it as
    # QUOTE_MINIMAL strips the quotes from every field that does not strictly need
    # them, which verify.py cannot see (it compares parsed cells, and both files
    # parse identically) while a fixed-position or naive-split consumer sees a
    # different file entirely.
    if '"' in text:
        fields = next(csv.reader([probe], delimiter=delimiter, quotechar='"'), [])
        quoting = (csv.QUOTE_ALL if fields and probe.strip() == delimiter.join(f'"{f}"' for f in fields)
                   else csv.QUOTE_MINIMAL)
    else:
        quoting = csv.QUOTE_NONE
    return {
        "encoding": encoding, "bom": bom, "newline": newline, "delimiter": delimiter,
        "text": text, "final_newline": final_newline, "quoting": quoting,
    }


def redact_delimited(inp: Path, outp: Path, plan: dict, salt: str) -> dict:
    meta = sniff_text(inp, plan.get("header_row", 0))
    delim = plan.get("delimiter") or meta["delimiter"]
    reader = csv.reader(io.StringIO(meta["text"]), delimiter=delim,
                        quoting=csv.QUOTE_MINIMAL, quotechar='"')
    rows = list(reader)
    if not rows:
        shutil.copy2(inp, outp)
        return {"rows": 0}

    header_row = plan.get("header_row", 0)
    header = rows[header_row]
    red = Redactor(plan, salt)

    name_cols = [k for k, e in red.columns.items() if e.get("type") == "name"]
    red.collect_name_words(
        [{header[j].strip().lower(): r[j] for j in range(min(len(header), len(r)))}
         for r in rows[header_row + 1:] if r],
        name_cols)

    out_rows = list(rows)
    for i in range(header_row + 1, len(rows)):
        row = rows[i]
        if not row or all(c == "" for c in row):
            continue  # blank line: keep as-is, it is part of the shape
        rowdict = {header[j].strip().lower(): row[j] for j in range(min(len(header), len(row)))}
        subject = red._row_identity(rowdict)
        new_row = list(row)
        for j in range(len(row)):
            if j < len(header):
                new_row[j] = red.redact_cell(header[j], row[j], subject)
        out_rows[i] = new_row

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=delim, quotechar='"',
                        quoting=meta["quoting"], lineterminator=meta["newline"])
    writer.writerows(out_rows)
    text = buf.getvalue()
    if not meta["final_newline"] and text.endswith(meta["newline"]):
        text = text[: -len(meta["newline"])]
    data = text.encode(meta["encoding"], errors="replace")
    if meta["bom"]:
        data = b"\xef\xbb\xbf" + data
    outp.write_bytes(data)
    return {"rows": len(rows) - header_row - 1, "cells": red.stats,
            "subjects": len(red._personas), "encoding": meta["encoding"], "delimiter": delim}


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def redact_excel(inp: Path, outp: Path, plan: dict, salt: str) -> dict:
    """Rewrite cell values in place on a copy, so everything else survives.

    Copying the file first and editing only the cells that need it is what keeps
    tab order, column widths, number formats, freeze panes, formulas and workbook
    properties intact. Rebuilding a workbook from scratch loses all of that.
    """
    import openpyxl

    shutil.copy2(inp, outp)
    # keep_vba preserves macros in .xlsm; without it openpyxl silently drops them
    # and the "everything else survives" promise quietly stops holding.
    wb = openpyxl.load_workbook(outp, keep_vba=inp.suffix.lower() == ".xlsm")
    red = Redactor(plan, salt)
    header_row = plan.get("header_row", 0) + 1  # openpyxl is 1-indexed
    sheet_plans = plan.get("sheets") or {}
    touched = {}

    targets = []
    for ws in wb.worksheets:
        sp = sheet_plans.get(ws.title)
        if sp is not None and sp.get("skip"):
            continue
        hrow = (sp or {}).get("header_row", header_row - 1) + 1 if sp else header_row
        if ws.max_row < hrow:
            continue
        targets.append((ws, hrow, [(c.value if c.value is not None else "") for c in ws[hrow]]))

    # Collect every real name word across the whole workbook before generating any
    # replacement. Per sheet, the first sheet is fully redacted before the second
    # sheet's names are even read, so a fake generated for sheet 1 can land on a
    # real name that appears only on sheet 2. That is coincidence rather than
    # disclosure, but it is indistinguishable from a leak to anyone auditing the
    # output -- which is the whole reason the avoidance pass exists.
    name_cols = [k for k, e in red.columns.items() if e.get("type") == "name"]
    if name_cols:
        for ws, hrow, header in targets:
            red.collect_name_words(
                [{str(n).strip().lower(): ("" if ws.cell(row=r, column=j + 1).value is None
                                           else str(ws.cell(row=r, column=j + 1).value))
                  for j, n in enumerate(header)}
                 for r in range(hrow + 1, ws.max_row + 1)],
                name_cols)

    for ws, hrow, header in targets:
        count = 0
        for r in range(hrow + 1, ws.max_row + 1):
            rowdict = {}
            for j, name in enumerate(header):
                cell = ws.cell(row=r, column=j + 1)
                rowdict[str(name).strip().lower()] = "" if cell.value is None else str(cell.value)
            if all(v == "" for v in rowdict.values()):
                continue
            subject = red._row_identity(rowdict)
            for j, name in enumerate(header):
                if not str(name).strip():
                    continue
                entry = red.columns.get(str(name).strip().lower())
                if entry is None:
                    continue
                cell = ws.cell(row=r, column=j + 1)
                if cell.value is None or isinstance(cell.value, str) and cell.value.startswith("="):
                    continue
                new = _redact_typed_cell(red, str(name), cell, subject, entry)
                if new is not None:
                    cell.value = new
                    count += 1
        touched[ws.title] = count

    props = _scrub_workbook_properties(wb)
    wb.save(outp)
    return {"sheets": touched, "cells": red.stats, "subjects": len(red._personas),
            "properties_cleared": props}


# Document properties travel with a workbook and survive the copy-and-edit round
# trip untouched. On a file that came from a client, lastModifiedBy is routinely a
# real person's name and title/subject sometimes names a patient -- PHI sitting
# outside every cell the column plan can see.
_SCRUBBED_PROPERTIES = ("creator", "lastModifiedBy", "title", "subject",
                        "description", "keywords", "category", "company", "manager")


def _scrub_workbook_properties(wb) -> list[str]:
    """Blank identifying document properties, returning the ones that held a value."""
    cleared = []
    for field in _SCRUBBED_PROPERTIES:
        if getattr(wb.properties, field, None):
            cleared.append(field)
            setattr(wb.properties, field, None)
    return cleared


def _redact_typed_cell(red: Redactor, col: str, cell, subject: str, entry: dict):
    """Keep the cell's Python type, so Excel's own formatting still applies.

    A real date cell must stay a date object and a numeric id must stay numeric;
    turning either into a string changes how Excel renders and sorts it, which is
    shape drift even though the characters look right.
    """
    if entry.get("type") == "keep":
        # Checked here as well as in redact_cell: date and numeric cells are handled
        # below without ever reaching redact_cell, so a reviewed-and-cleared column
        # of real date cells would otherwise still get shifted.
        return None

    val = cell.value
    if isinstance(val, (datetime.datetime, datetime.date)):
        off = red.date_offset(subject)
        shifted = val + datetime.timedelta(days=off)
        red.stats["date"] = red.stats.get("date", 0) + 1
        return shifted
    if isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        # Numbers must come back as numbers. Excel sorts, formats and aggregates by
        # cell type, so a numeric id that returns as text changes how the sheet
        # behaves even when the characters match.
        text = str(int(val)) if float(val).is_integer() else str(val)
        new = red.redact_cell(col, text, subject)
        try:
            return int(new) if float(val).is_integer() else float(new)
        except (ValueError, TypeError):
            return new
    return red.redact_cell(col, str(val), subject)


def output_name(inp: Path, suffix: str = "_REDACTED") -> Path:
    """Keep the stem and extension; insert a marker so nobody confuses the two.

    The name stays recognisable because matching filenames is how these files get
    routed and how a human finds the right one, but an unmistakable marker is
    worth more than perfect mimicry -- a redacted file that looks byte-identical
    in a directory listing is an invitation to hand the wrong one to a client.
    """
    return inp.with_name(f"{inp.stem}{suffix}{inp.suffix}")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--plan", required=True, type=Path)
    ap.add_argument("--in", dest="inp", required=True, type=Path)
    ap.add_argument("--out", dest="outp", type=Path)
    ap.add_argument("--salt", default=None,
                    help=f"defaults to ${SALT_ENV}, else a fresh random salt per run")
    args = ap.parse_args()

    plan = json.loads(args.plan.read_text())
    # A fresh random salt by default. Every fake value derives from
    # sha256(salt || column || real value), so a salt that ships in this file is a
    # re-identification key: anyone holding the redacted output and this repo can
    # brute-force the small identifier spaces (a 7-digit MRN is 10M guesses) and
    # confirm a hit because the same subject key also predicts the fake name and
    # date shift. That is precisely the crosswalk this skill refuses to write.
    # Reproducibility is opt-in, and the salt chosen for it must be kept secret.
    salt = args.salt or os.environ.get(SALT_ENV) or secrets.token_hex(16)
    outp = args.outp or output_name(args.inp, plan.get("output_suffix", "_REDACTED"))
    outp.parent.mkdir(parents=True, exist_ok=True)

    ext = args.inp.suffix.lower()
    if ext == ".xls":
        # openpyxl reads the OOXML formats only; handing it a legacy BIFF .xls
        # raises deep in the zip reader with a message that reads like a corrupt
        # file rather than an unsupported one.
        print("error: legacy .xls is not supported -- re-save as .xlsx first", file=sys.stderr)
        return 2
    if ext in {".xlsx", ".xlsm"}:
        info = redact_excel(args.inp, outp, plan, salt)
    else:
        info = redact_delimited(args.inp, outp, plan, salt)

    print(json.dumps({"output": str(outp), **info}, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
