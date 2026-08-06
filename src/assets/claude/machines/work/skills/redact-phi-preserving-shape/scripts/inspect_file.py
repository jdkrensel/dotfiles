"""Profile a file and scaffold a redaction plan for review.

Usage:
    uv run --with openpyxl python inspect_file.py FILE [--rows 8] [--emit-plan plan.json]

This exists to make the judgement step cheap and well-informed rather than to
replace it. It reports each column's structure -- distinct-value count, character
masks, name templates, detected dates -- and proposes a type where the evidence is
strong.

Treat the proposal as a draft. Header names lie: a column called "ID" might be a
patient identifier or a procedure code, and only the values and surrounding
context tell you which. The value samples printed here are the evidence you need
to make that call. Anything the script is unsure of is emitted with
"type": "REVIEW" so it cannot be applied by accident.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import re
import sys
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from redact import sniff_text  # noqa: E402
from shapes import mask_of, name_template, parse_date  # noqa: E402

# Header-name evidence. Matched against a normalized header; a hit here is a
# strong hint but the value evidence below can and should override it.
HINTS: list[tuple[str, str]] = [
    (r"\bssn\b|social.?sec", "ssn"),
    (r"\b(dob|birth)\b|date.?of.?birth|birthdate", "date"),
    (r"\bdod\b|date.?of.?death|deceased.?date", "date"),
    (r"\b(dos|admit|discharge|service|encounter|visit|surgery|appt|appointment|scheduled|proc(edure)?)\b.*date|"
     r"date.*\b(dos|admit|discharge|service|encounter|visit|surgery)\b|\bdos\b", "date"),
    (r"^(pat(ient)? )?name$|pat(ient)?.?name|\bfull.?name\b|last.?first", "name"),
    (r"first.?name|^f.?name$|\bgiven\b", "name"),
    (r"last.?name|^l.?name$|surname|family.?name", "name"),
    (r"middle.?name|^m.?name$|middle.?init|\bmi\b", "name"),
    (r"\b(mrn|mr#|medical.?record)\b|\bpat(ient)?.?(id|num)", "id"),
    (r"phone|\btel\b|mobile|cell|fax", "phone"),
    (r"e.?mail", "email"),
    (r"\bzip\b|postal", "zip"),
    (r"\bcity\b|municipal", "city"),
    (r"address|\bstreet\b|addr\d?|\bline[12]\b", "address"),
    (r"\bage\b", "age"),
    (r"subscriber|member.?(id|num)|policy|beneficiary|insurance.?(id|num)|health.?plan", "id"),
    (r"account.?(num|no|#)|\bacct\b|\bguarantor\b|\bfin\b|\bhar\b|billing.?(num|id)", "id"),
    (r"\bcsn\b|encounter.?(csn|id|num)|visit.?(id|num)|\bcase.?(id|num)\b", "id"),
    (r"licen[cs]e|\bnpi\b|certificate", "id"),
    (r"\bvin\b|licen[cs]e.?plate|\bplate\b", "id"),
    (r"serial|device.?(id|num)|implant.?(id|serial)|\blot\b|\budi\b", "id"),
    (r"\burl\b|\bweb\b|portal.?link|\blink\b", "url"),
    (r"ip.?addr|\bipv?[46]?.?address\b", "ip"),
    (r"\bguid\b|\buuid\b", "id"),
]

# Columns that are clinical or administrative vocabulary, not identifiers. Keeping
# these intact is what makes a redacted file useful, so they are proposed as keep.
KEEP_HINTS = [
    r"\bcpt\b", r"icd", r"\bhcpcs\b", r"\bdrg\b", r"diagnos", r"proc.?(name|desc)",
    r"payor|payer|insurance.?(name|plan)|plan.?name", r"lateral", r"\bsex\b|gender",
    r"\brace\b|ethnic", r"\blanguage\b", r"\bstate\b", r"\bcountry\b", r"consent",
    r"\bstatus\b", r"hospital|facility|clinic|location|department|\bunit\b|\bward\b|\bsite\b",
    r"surgeon|provider|physician|\bdoctor\b|\bmd\b|\borp\b|attending|referring|operator",
    r"\btype\b", r"\bclass\b", r"\bcategory\b", r"quantity|\bqty\b|amount|\bcharge\b|\bcost\b|price",
    r"\bunits?\b", r"modifier", r"revenue.?code", r"\bpos\b", r"\bnote", r"comment", r"\bdesc",
]


def classify(header: str, values: list[str]) -> tuple[str, str]:
    """Propose a type for a column, with the reason. Returns (type, rationale)."""
    h = re.sub(r"[^a-z0-9]+", " ", header.lower()).strip()
    nonempty = [v for v in values if v.strip()]

    if not nonempty:
        return "keep", "column is empty in sample"

    # Value evidence first for the unambiguous formats -- these beat any header name.
    if all(re.fullmatch(r"\d{3}-?\d{2}-?\d{4}", v.strip()) for v in nonempty):
        return "ssn", "every value matches SSN format"
    if all("@" in v and "." in v.split("@")[-1] for v in nonempty):
        return "email", "every value looks like an email address"
    if all(re.fullmatch(r"(\d{1,3}\.){3}\d{1,3}", v.strip()) for v in nonempty):
        return "ip", "every value is a dotted-quad IP"
    if all(re.match(r"https?://|www\.", v.strip(), re.I) for v in nonempty):
        return "url", "every value is a URL"

    dates = sum(1 for v in nonempty if parse_date(v.strip()))
    date_ratio = dates / len(nonempty)

    keep_hit = next((p for p in KEEP_HINTS if re.search(p, h)), None)
    hint_type = next((t for p, t in HINTS if re.search(p, h)), None)

    # A date-shaped column named for a person-linked event is PHI; a date-shaped
    # column that is clinical vocabulary (e.g. a code's effective date) usually is not,
    # so the header hint has to agree before dates are shifted.
    if date_ratio > 0.9:
        if hint_type == "date":
            return "date", f"header suggests a person-linked date and {dates}/{len(nonempty)} values parse as dates"
        if keep_hit:
            return "REVIEW", f"values parse as dates but header matches keep-pattern /{keep_hit}/ -- is this tied to a person?"
        return "REVIEW", f"{dates}/{len(nonempty)} values parse as dates but header is unclear -- person-linked?"

    if keep_hit and hint_type is None:
        return "keep", f"header matches non-identifier pattern /{keep_hit}/"

    if hint_type == "name":
        tmpls = Counter(name_template(v.strip()) for v in nonempty)
        looks_like_name = all(re.fullmatch(r"[A-Za-z ,.'\-]+", v.strip()) for v in nonempty)
        if looks_like_name:
            part = ("first" if re.search(r"first|given|^f_?name", h) else
                    "last" if re.search(r"last|surname|family|^l_?name", h) else
                    "middle" if re.search(r"middle|^m_?name", h) else None)
            order = "last_first" if any("," in v for v in nonempty) else "first_last"
            return "name", f"name-shaped values, {len(tmpls)} distinct templates, order={order}, part={part}"
        return "REVIEW", "header suggests a name but values are not name-shaped"

    if hint_type:
        return hint_type, f"header matches /{next(p for p, t in HINTS if re.search(p, h) and t == hint_type)}/"

    # No header signal. Decide from cardinality: a high-cardinality opaque token is
    # almost certainly an identifier; a low-cardinality one is a category.
    distinct = len(set(v.strip() for v in nonempty))
    ratio = distinct / len(nonempty)
    idish = sum(1 for v in nonempty if re.fullmatch(r"[A-Za-z0-9\-]{5,}", v.strip()) and any(c.isdigit() for c in v))
    if ratio > 0.5 and idish / len(nonempty) > 0.8:
        return "REVIEW", f"high-cardinality opaque tokens ({distinct} distinct) -- likely an identifier, confirm"
    if ratio < 0.2:
        return "keep", f"low cardinality ({distinct} distinct values) -- looks categorical"
    return "REVIEW", f"no strong signal ({distinct} distinct, {ratio:.0%} unique)"


def profile_columns(header: list[str], rows: list[list[str]], sample: int) -> list[dict]:
    out = []
    for j, name in enumerate(header):
        vals = [r[j] for r in rows if j < len(r)]
        nonempty = [v for v in vals if v.strip()]
        ctype, why = classify(str(name), vals)
        masks = Counter(mask_of(v.strip()) for v in nonempty).most_common(3)
        tmpls = Counter(name_template(v.strip()) for v in nonempty).most_common(3) if ctype == "name" else []
        out.append({
            "index": j, "name": name, "proposed_type": ctype, "rationale": why,
            "filled": f"{len(nonempty)}/{len(vals)}",
            "distinct": len(set(v.strip() for v in nonempty)),
            "samples": [v for v in nonempty[:sample]],
            "masks": masks, "templates": tmpls,
        })
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("path", type=Path)
    ap.add_argument("--rows", type=int, default=6, help="sample values to show per column")
    ap.add_argument("--header-row", type=int, default=0)
    ap.add_argument("--emit-plan", type=Path, help="write a draft plan for editing")
    args = ap.parse_args()

    ext = args.path.suffix.lower()
    sheets: dict[str, list[dict]] = {}
    meta: dict = {}

    if ext in {".xlsx", ".xlsm", ".xls"}:
        import openpyxl
        wb = openpyxl.load_workbook(args.path, data_only=True)
        meta = {"format": "excel", "sheets": wb.sheetnames}
        for ws in wb.worksheets:
            rows = [[("" if c.value is None else str(c.value)) for c in row]
                    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 400))]
            if len(rows) <= args.header_row:
                sheets[ws.title] = []
                continue
            sheets[ws.title] = profile_columns(rows[args.header_row], rows[args.header_row + 1:], args.rows)
    else:
        m = sniff_text(args.path, args.header_row)
        rows = list(csv.reader(io.StringIO(m["text"]), delimiter=m["delimiter"], quotechar='"'))
        rows = [r for r in rows if r and any(c.strip() for c in r)]
        meta = {"format": "delimited", "delimiter": m["delimiter"], "encoding": m["encoding"],
                "newline": repr(m["newline"]), "bom": m["bom"], "rows": len(rows) - 1}
        sheets["(single)"] = profile_columns(rows[args.header_row], rows[args.header_row + 1:401], args.rows)

    print(json.dumps(meta, indent=2))
    for sheet, cols in sheets.items():
        print(f"\n=== {sheet} ===")
        for c in cols:
            flag = "  <-- REVIEW" if c["proposed_type"] == "REVIEW" else ""
            print(f"\n[{c['index']:>3}] {c['name']!r}  ->  {c['proposed_type']}{flag}")
            print(f"      why: {c['rationale']}")
            print(f"      filled {c['filled']}, {c['distinct']} distinct")
            for v in c["samples"]:
                print(f"      e.g. {v!r}")
            if c["masks"]:
                print(f"      masks: {', '.join(f'{m}({n})' for m, n in c['masks'])}")
            if c["templates"]:
                print(f"      templates: {', '.join(f'{t}({n})' for t, n in c['templates'])}")

    if args.emit_plan:
        plan = {
            "_comment": "Review every entry. 'REVIEW' types must be resolved to a real type or to 'keep' before use. 'keep' means reviewed and judged non-PHI: the column is preserved byte-for-byte.",
            "header_row": args.header_row,
            "identity_key": [],
            "date_offset_range": [-540, 540],
            "output_suffix": "_REDACTED",
            "columns": [],
        }
        if meta.get("format") == "delimited":
            plan["delimiter"] = meta["delimiter"]
        # Columns judged non-PHI are emitted as explicit "keep" entries rather than
        # dropped. verify.py re-profiles the file independently and fails on any
        # unlisted column that looks like PHI, so a silently-omitted column would
        # come back as a failure the user has to re-decide on every run. Listing
        # every column also makes the plan a readable record of what was considered.
        seen = set()
        for cols in sheets.values():
            for c in cols:
                if c["name"] in seen:
                    continue
                seen.add(c["name"])
                entry = {"name": c["name"], "type": c["proposed_type"], "_why": c["rationale"]}
                if c["proposed_type"] == "name":
                    if "part=" in c["rationale"]:
                        part = c["rationale"].split("part=")[1].split(",")[0].strip()
                        if part != "None":
                            entry["name_part"] = part
                    if "order=" in c["rationale"]:
                        entry["name_order"] = c["rationale"].split("order=")[1].split(",")[0].strip()
                plan["columns"].append(entry)
        args.emit_plan.write_text(json.dumps(plan, indent=2) + "\n")
        review = [c["name"] for cols in sheets.values() for c in cols if c["proposed_type"] == "REVIEW"]
        print(f"\nDraft plan -> {args.emit_plan}")
        print("identity_key is empty -- set it to the column(s) identifying a patient (e.g. the MRN).")
        if review:
            print(f"{len(review)} column(s) need a decision: {', '.join(map(str, review))}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
