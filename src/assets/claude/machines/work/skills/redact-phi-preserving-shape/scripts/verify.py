"""Verify a redacted file: same shape, no surviving PHI.

Usage:
    uv run --with openpyxl python verify.py --original REAL --redacted FAKE [--plan plan.json]

Two independent questions, and both have to be answered before a redacted file
is trustworthy:

  1. Is the shape identical? Row and column counts, headers, tab names,
     delimiter, encoding, per-cell character masks, name templates, date formats.
     Shape drift means the file no longer exercises the same code paths.

  2. Did any real value survive? Every PHI-column value in the original is
     checked for presence anywhere in the redacted output -- a row skipped, a
     value that slipped through unchanged, a name that migrated into a free-text
     column.

That second check has a blind spot worth naming, because it shaped the design: it
can only hunt for values the plan told it were PHI. A column the plan never
mentions is neither redacted nor searched for, so on its own the scan would report
a clean file while an overlooked column sat there in full. Verification therefore
forms its own opinion, independent of the plan, and fails if:

  - a column absent from the plan looks like PHI to the same classifier
    inspect_file.py uses (resolve a false positive by recording the decision as
    `{"type": "keep"}`, not by loosening the check); or
  - a plan column matches no header in the file, which is how a header typo or a
    renamed column silently disables an entry.

Exit code is non-zero if anything fails, so this can gate a workflow.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import re
import sys
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from inspect_file import classify  # noqa: E402
from redact import _SCRUBBED_PROPERTIES, sniff_text  # noqa: E402
from shapes import NAME_SUFFIXES, mask_of, name_template, parse_date  # noqa: E402


class Report:
    def __init__(self) -> None:
        self.checks: list[dict] = []

    def add(self, name: str, passed: bool, detail: str = "", advisory: bool = False) -> None:
        """Record a check. Advisory ones report but never change the exit code.

        Only findings that mean "do not ship this file" may gate the exit code. A
        naming convention is worth mentioning and worth nothing as a gate -- and a
        verifier that exits non-zero over cosmetics teaches its reader that red is
        sometimes fine, which is the one lesson this tool cannot afford to teach.
        """
        self.checks.append({"check": name, "passed": bool(passed),
                            "detail": detail, "advisory": advisory})

    @property
    def ok(self) -> bool:
        return all(c["passed"] for c in self.checks if not c["advisory"])

    def render(self) -> str:
        lines = []
        for c in self.checks:
            mark = "PASS" if c["passed"] else ("WARN" if c["advisory"] else "FAIL")
            lines.append(f"[{mark}] {c['check']}" + (f" -- {c['detail']}" if c["detail"] else ""))
        lines.append("")
        gating = [c for c in self.checks if not c["advisory"]]
        failed = sum(1 for c in gating if not c["passed"])
        warned = sum(1 for c in self.checks if c["advisory"] and not c["passed"])
        lines.append(f"{len(gating) - failed}/{len(gating)} checks passed"
                     + ("" if not failed else f"  ({failed} FAILED)")
                     + ("" if not warned else f"  ({warned} warning)"))
        return "\n".join(lines)


def _read_delimited(path: Path, plan: dict) -> tuple[list[list[str]], dict]:
    """Parse exactly the way redact.py did, or the comparison is meaningless.

    Both the header row and an explicit delimiter override come from the plan. A
    verifier that sniffs independently can split the two files into different
    column grids and then report confidently on a comparison it never made.
    """
    meta = sniff_text(path, plan.get("header_row", 0))
    delim = plan.get("delimiter") or meta["delimiter"]
    meta["delimiter"] = delim
    rows = list(csv.reader(io.StringIO(meta["text"]), delimiter=delim, quotechar='"'))
    return rows, meta


def verify_delimited(orig: Path, red: Path, plan: dict, rep: Report) -> None:
    o_rows, o_meta = _read_delimited(orig, plan)
    r_rows, r_meta = _read_delimited(red, plan)

    rep.add("encoding preserved", o_meta["encoding"] == r_meta["encoding"],
            f"{o_meta['encoding']} vs {r_meta['encoding']}")
    rep.add("delimiter preserved", o_meta["delimiter"] == r_meta["delimiter"],
            f"{o_meta['delimiter']!r} vs {r_meta['delimiter']!r}")
    rep.add("newline style preserved", o_meta["newline"] == r_meta["newline"],
            f"{o_meta['newline']!r} vs {r_meta['newline']!r}")
    rep.add("BOM preserved", o_meta["bom"] == r_meta["bom"])
    rep.add("row count identical", len(o_rows) == len(r_rows), f"{len(o_rows)} vs {len(r_rows)}")

    hr = plan.get("header_row", 0)
    if o_rows and r_rows:
        rep.add("header values identical", o_rows[hr] == r_rows[hr],
                "" if o_rows[hr] == r_rows[hr] else f"{o_rows[hr]} vs {r_rows[hr]}")

    widths_match = all(len(a) == len(b) for a, b in zip(o_rows, r_rows))
    bad = next((i for i, (a, b) in enumerate(zip(o_rows, r_rows)) if len(a) != len(b)), None)
    rep.add("every row has same column count", widths_match,
            "" if widths_match else f"first mismatch at line {bad + 1}")

    header = [h.strip().lower() for h in o_rows[hr]] if o_rows else []
    raw_header = o_rows[hr] if o_rows else []
    phi_cols, named = _plan_cols(plan)
    _verify_plan_resolved(header, plan, named, rep)
    _verify_unplanned_columns(o_rows, header, raw_header, named, hr, rep)
    _verify_cells(o_rows, r_rows, header, phi_cols, hr, rep)
    _verify_leaks(o_rows, r_rows, header, phi_cols, hr, rep)


def _plan_cols(plan: dict) -> tuple[dict[str, dict], set[str]]:
    """Split the plan into columns to redact and columns merely acknowledged.

    A `"type": "keep"` entry records a reviewed decision that a column is not PHI.
    It is not redacted, but it counts as considered -- which is what separates it
    from a column nobody ever looked at.
    """
    cols: dict[str, dict] = {}
    named: set[str] = set()
    for e in plan.get("columns", []):
        for n in [e["name"]] + list(e.get("aliases", [])):
            key = n.strip().lower()
            named.add(key)
            if e.get("type") != "keep":
                cols[key] = e
    return cols, named


def _verify_plan_resolved(header: list[str], plan: dict, named: set[str], rep: Report) -> None:
    """Fail when a plan column matches no header in the file.

    A plan is allowed to cover a family of similar files, so an unmatched entry was
    treated as harmless. It is not: one renamed column, one stray non-breaking
    space in a header, and the entry silently stops applying while every other
    check still passes. Naming the misses is the only way the user finds out.
    """
    present = set(header)
    missing = sorted(
        e["name"] for e in plan.get("columns", [])
        if not ({e["name"].strip().lower()} | {a.strip().lower() for a in e.get("aliases", [])}) & present
    )
    rep.add("every plan column matched a header in the file", not missing,
            f"unmatched: {_sample(missing, 5)}" if missing else f"{len(named)} names/aliases in plan")


def _verify_unplanned_columns(o_rows, header, raw_header, named: set[str], hr: int, rep: Report) -> None:
    """Judge PHI-ness from the file itself, with no reference to the plan.

    This is the check that makes the leak scan below meaningful. That scan can only
    hunt for values it already believes are PHI, and it learns which those are from
    the plan -- so a column the plan never mentions is neither redacted nor
    searched for, and the exact failure this tool exists to catch is the one it is
    structurally blind to. Re-profiling the original with the same classifier
    inspect_file.py uses restores an opinion the plan cannot suppress.

    Resolve a false positive by recording the decision as `{"type": "keep"}` rather
    than by loosening this check: the point is that every column was considered.
    """
    suspects = _phi_shaped_unplanned(header, raw_header, named,
                                     lambda j: [o_rows[i][j] for i in range(hr + 1, len(o_rows))
                                                if j < len(o_rows[i])])
    rep.add("no PHI-shaped column is missing from the plan", not suspects, _sample(suspects, 4))


# inspect_file profiles the first 400 data rows and drops every column it judges
# "keep" from the draft plan. Classifying the whole column here would let verify
# reach a different verdict on the same data -- flagging a column the tool itself
# told the user to omit -- so both sides look at the same slice.
_CLASSIFY_SAMPLE = 400


def _phi_shaped_unplanned(header, raw_header, named: set[str], values_at) -> list[str]:
    """Names of unplanned columns that the classifier judges to hold PHI."""
    suspects = []
    for j, col in enumerate(header):
        if col in named:
            continue
        values = values_at(j)[:_CLASSIFY_SAMPLE]
        if not any(v.strip() for v in values):
            continue  # empty throughout: nothing to leak
        ctype, why = classify(str(raw_header[j]), values)
        if ctype == "keep":
            continue
        # Unplanned means copied through verbatim, so a PHI-shaped column here is
        # not a risk of a leak -- it is the leak, already written to the output.
        suspects.append(f"'{raw_header[j]}' looks like {ctype} -- {why}")
    return suspects


def _verify_cells(o_rows, r_rows, header, phi_cols, hr, rep: Report) -> None:
    """Compare shape cell by cell and confirm PHI cells actually changed."""
    mask_mismatch: list[str] = []
    unchanged: list[str] = []
    nonphi_changed: list[str] = []
    template_mismatch: list[str] = []
    datefmt_mismatch: list[str] = []
    changed_count = 0
    phi_cell_count = 0

    for i in range(hr + 1, min(len(o_rows), len(r_rows))):
        o_row, r_row = o_rows[i], r_rows[i]
        for j in range(min(len(o_row), len(r_row), len(header))):
            ov, rv = o_row[j], r_row[j]
            col = header[j]
            entry = phi_cols.get(col)
            if entry is None:
                if ov != rv:
                    nonphi_changed.append(f"line {i + 1} col '{header[j]}': {ov!r} -> {rv!r}")
                continue

            if ov.strip() == "":
                if rv.strip() != "":
                    mask_mismatch.append(f"line {i + 1} col '{header[j]}': empty became {rv!r}")
                continue
            phi_cell_count += 1  # non-empty only: an empty cell has nothing to rewrite

            kind_now = entry.get("type", "id")
            if kind_now in _LOOSE_MASK_TYPES:
                if _loose_shape(ov) != _loose_shape(rv):
                    mask_mismatch.append(
                        f"line {i + 1} col '{header[j]}': structure changed ({ov!r} -> {rv!r})")
            elif mask_of(ov) != mask_of(rv):
                mask_mismatch.append(
                    f"line {i + 1} col '{header[j]}': mask {mask_of(ov)!r} -> {mask_of(rv)!r} ({ov!r} -> {rv!r})")

            if ov == rv:
                unchanged.append(f"line {i + 1} col '{header[j]}': {ov!r}")
            else:
                changed_count += 1

            kind = entry.get("type", "id")
            if kind == "name" and name_template(ov) != name_template(rv):
                template_mismatch.append(
                    f"line {i + 1} col '{header[j]}': {name_template(ov)!r} -> {name_template(rv)!r} ({ov!r} -> {rv!r})")
            if kind == "date":
                po, pr = parse_date(ov.strip()), parse_date(rv.strip())
                if po and (pr is None or _fmt_sig(po) != _fmt_sig(pr)):
                    datefmt_mismatch.append(f"line {i + 1} col '{header[j]}': {ov!r} -> {rv!r}")

    rep.add("non-PHI columns byte-identical", not nonphi_changed,
            _sample(nonphi_changed))
    rep.add("PHI cell character masks preserved", not mask_mismatch, _sample(mask_mismatch))
    rep.add("no PHI cell left unchanged", not unchanged, _sample(unchanged))
    rep.add("name templates preserved per row", not template_mismatch, _sample(template_mismatch))
    rep.add("date formats preserved", not datefmt_mismatch, _sample(datefmt_mismatch))
    # phi_cell_count now counts non-empty cells only, so "nothing to rewrite" is a
    # genuine state (a header-only file, a column blank throughout) rather than the
    # blanket excuse it used to be. The cases this once waved through -- a header
    # typo, a delimiter sniffed off a banner row, a renamed column, all of which
    # produced a fully unredacted file with a green report -- are caught up front
    # by "every plan column matched a header in the file".
    rep.add("PHI cells were rewritten", changed_count > 0 or phi_cell_count == 0,
            f"{changed_count}/{phi_cell_count} non-empty PHI cells changed")


def _loose_shape(v: str) -> tuple:
    """Structural signature for values whose exact length is allowed to change.

    Captures what a parser cares about -- token count, the presence of an @, and
    which tokens are numeric -- without demanding character-for-character length.
    """
    toks = v.split()
    return (len(toks), "@" in v, v.count("."), v.count(","),
            tuple(t.isdigit() for t in toks), tuple(len(t) if t.isdigit() else 0 for t in toks))


def _fmt_sig(p: dict) -> tuple:
    if p["kind"] == "numeric":
        return ("numeric", p["order"], p["widths"], p["seps"])
    if p["kind"] == "text":
        return ("text", p["sep"], p["ywidth"], p["dwidth"], len(p["mon_style"]))
    return ("compact",)


# Types whose replacement is deliberately not mask-exact, and why:
#   email   -- the domain is forced to a reserved one (example.com) so the address
#              provably cannot reach a real mailbox. A shape-preserved domain would
#              be a live route to a real person, which no amount of fidelity earns.
#   city    -- drawn from a fixed pool of invented place names; padding one to an
#              exact length would produce visible gibberish in a human-read column.
#   address -- street names likewise come from a pool.
# For these, length may differ; structure (the @, the token count) is still checked.
_LOOSE_MASK_TYPES = {"email", "city", "address", "url"}

# Types whose values live in a small space, where a generated value colliding with
# some *other* subject's real value is arithmetic rather than disclosure. A date has
# only ~36k plausible spellings and an age ~90, so across a few hundred rows
# collisions are certain. For these, "did it change?" is the meaningful test and is
# already covered cell-by-cell; a global substring scan only produces noise.
_SMALL_SPACE_TYPES = {"date", "age", "zip", "city"}


def _verify_leaks(o_rows, r_rows, header, phi_cols, hr, rep: Report) -> None:
    """Search the redacted file for any original PHI value that survived.

    The scan is global rather than per-column, because a name that migrated into a
    free-text column is still a breach. Short values are skipped since 1-2
    character tokens collide by chance.

    Scope note: the haystack is built from plan columns, so this catches values
    that escaped their column -- not columns the plan never named. That gap is
    _verify_unplanned_columns' job, and neither check substitutes for the other.
    """
    whole: set[str] = set()      # complete original values
    fragments: set[str] = set()  # individual name words
    for i in range(hr + 1, len(o_rows)):
        for j in range(min(len(o_rows[i]), len(header))):
            entry = phi_cols.get(header[j])
            if entry is None or entry.get("type") in _SMALL_SPACE_TYPES:
                continue
            v = o_rows[i][j].strip()
            if len(v) >= 3:
                whole.add(v)
                # A leak can be partial -- the surname surviving while the given name
                # changed -- so name words are tracked individually too.
                if entry.get("type") == "name":
                    for part in re.split(r"[,\s]+", v):
                        bare = part.rstrip(".")
                        if len(bare) >= 4 and bare.upper() not in _STRUCTURAL:
                            fragments.add(bare)

    # Two haystacks, because the columns differ in what a hit means. In a redacted
    # column, any trace of an original value is a defect. In a deliberately preserved
    # column, only a *whole* original value is: single name words legitimately recur
    # there (a surgeon sharing a given name with a patient's middle name), and
    # flagging those buries real findings under noise.
    phi_idx = [j for j in range(len(header)) if header[j] in phi_cols]
    keep_idx = [j for j in range(len(header)) if header[j] not in phi_cols]
    phi_hay = "\n".join("\x1f".join(r[j] for j in phi_idx if j < len(r)) for r in r_rows[hr + 1:])
    # Rows at or above the header go in the preserved haystack. redact.py starts
    # below the header, so a banner line ("Report for A. Test -- 2026-01-02") is
    # copied through verbatim; searching only the data rows would leave the one
    # part of the file nothing ever rewrites completely unexamined.
    keep_hay = "\n".join(
        ["\x1f".join(r) for r in r_rows[:hr + 1]]
        + ["\x1f".join(r[j] for j in keep_idx if j < len(r)) for r in r_rows[hr + 1:]])

    # Whole values are matched as substrings, since a real value embedded in a larger
    # cell is still a leak. Name fragments are matched on word boundaries instead:
    # "FRANCES" occurring inside the generated "FRANCESCA" is a coincidence of
    # spelling, not a surviving identifier, and treating it as one would train the
    # reader to ignore this check.
    frag_hits = {v for v in fragments
                 if re.search(rf"(?<![A-Za-z]){re.escape(v)}(?![A-Za-z])", phi_hay, re.I)}
    leaked = sorted({v for v in whole if v in phi_hay} | frag_hits
                    | {v for v in whole if v in keep_hay})
    rep.add("no original PHI value survives in output", not leaked,
            f"{len(leaked)} leaked: {_sample(leaked, 5)}" if leaked else
            f"checked {len(whole)} values and {len(fragments)} name components")

    # Dates and ages get the check that is actually meaningful for them: within a
    # row, the value must differ from the original.
    small = _small_space_survivors(o_rows, r_rows, header, phi_cols, hr)
    rep.add("date/age values changed in every row", not small, _sample(small))


_STRUCTURAL = {s.rstrip(".") for s in NAME_SUFFIXES}

_IDENTIFYING_PROPERTIES = _SCRUBBED_PROPERTIES


def _small_space_survivors(o_rows, r_rows, header, phi_cols, hr) -> list[str]:
    out = []
    for i in range(hr + 1, min(len(o_rows), len(r_rows))):
        for j in range(min(len(o_rows[i]), len(r_rows[i]), len(header))):
            entry = phi_cols.get(header[j])
            if entry is None or entry.get("type") not in _SMALL_SPACE_TYPES:
                continue
            ov, rv = o_rows[i][j].strip(), r_rows[i][j].strip()
            if ov and ov == rv:
                out.append(f"line {i + 1} col '{header[j]}': {ov!r} unchanged")
    return out


def _sample(items: list[str], n: int = 3) -> str:
    if not items:
        return ""
    shown = "; ".join(items[:n])
    return shown + (f" (+{len(items) - n} more)" if len(items) > n else "")


def verify_excel(orig: Path, red: Path, plan: dict, rep: Report) -> None:
    import openpyxl

    ob = openpyxl.load_workbook(orig)
    rb = openpyxl.load_workbook(red)

    rep.add("sheet count identical", len(ob.worksheets) == len(rb.worksheets),
            f"{len(ob.worksheets)} vs {len(rb.worksheets)}")
    rep.add("sheet names and order identical", ob.sheetnames == rb.sheetnames,
            f"{ob.sheetnames} vs {rb.sheetnames}")

    # Document properties survive the copy-and-edit round trip untouched, so they
    # are PHI the column plan structurally cannot see: lastModifiedBy on a client
    # workbook is routinely a real person and title sometimes names a patient.
    leftover = [f for f in _IDENTIFYING_PROPERTIES if getattr(rb.properties, f, None)]
    rep.add("workbook document properties scrubbed", not leftover,
            f"still set: {', '.join(leftover)}" if leftover else
            f"checked {len(_IDENTIFYING_PROPERTIES)} properties")

    phi_cols, named = _plan_cols(plan)
    sheet_plans = plan.get("sheets") or {}
    all_mask, all_unchanged, all_nonphi, all_tmpl = [], [], [], []
    r_texts: list[str] = []
    originals: set[str] = set()
    suspects: list[str] = []
    seen_headers: set[str] = set()
    changed = total = 0

    for o_ws, r_ws in zip(ob.worksheets, rb.worksheets):
        rep.add(f"[{o_ws.title}] dimensions identical",
                (o_ws.max_row, o_ws.max_column) == (r_ws.max_row, r_ws.max_column),
                f"{o_ws.max_row}x{o_ws.max_column} vs {r_ws.max_row}x{r_ws.max_column}")
        sp = sheet_plans.get(o_ws.title) or {}
        if sp.get("skip"):
            # redact_excel leaves a skipped sheet entirely alone, so verifying its
            # cells against the plan would flag every one of them as unredacted.
            continue
        hrow = sp.get("header_row", plan.get("header_row", 0)) + 1
        if o_ws.max_row < hrow:
            continue
        o_hdr = [c.value for c in o_ws[hrow]]
        r_hdr = [c.value for c in r_ws[hrow]]
        rep.add(f"[{o_ws.title}] header values identical", o_hdr == r_hdr,
                "" if o_hdr == r_hdr else f"{o_hdr} vs {r_hdr}")

        # The same plan-independent judgement the delimited path makes, per sheet.
        # Without it a workbook column the plan never mentioned is copied through
        # verbatim and nothing ever looks at it.
        sheet_keys = [str(n).strip().lower() if n is not None else "" for n in o_hdr]
        seen_headers.update(k for k in sheet_keys if k)
        suspects.extend(
            f"[{o_ws.title}] {s}" for s in _phi_shaped_unplanned(
                sheet_keys, [("" if n is None else n) for n in o_hdr], named,
                lambda j: ["" if o_ws.cell(row=r, column=j + 1).value is None
                           else str(o_ws.cell(row=r, column=j + 1).value)
                           for r in range(hrow + 1, min(o_ws.max_row, hrow + _CLASSIFY_SAMPLE) + 1)]))

        for r in range(hrow + 1, o_ws.max_row + 1):
            for j, name in enumerate(o_hdr):
                oc = o_ws.cell(row=r, column=j + 1)
                rc = r_ws.cell(row=r, column=j + 1)
                ov = "" if oc.value is None else str(oc.value)
                rv = "" if rc.value is None else str(rc.value)
                key = str(name).strip().lower() if name is not None else ""
                entry = phi_cols.get(key)
                if entry is None:
                    if ov != rv:
                        all_nonphi.append(f"[{o_ws.title}] {oc.coordinate}: {ov!r} -> {rv!r}")
                    continue
                if ov.strip() == "":
                    continue
                total += 1
                if type(oc.value) is not type(rc.value):
                    all_mask.append(f"[{o_ws.title}] {oc.coordinate}: type {type(oc.value).__name__} -> {type(rc.value).__name__}")
                elif isinstance(oc.value, str):
                    if entry.get("type") in _LOOSE_MASK_TYPES:
                        if _loose_shape(ov) != _loose_shape(rv):
                            all_mask.append(f"[{o_ws.title}] {oc.coordinate}: structure changed ({ov!r} -> {rv!r})")
                    elif mask_of(ov) != mask_of(rv):
                        all_mask.append(f"[{o_ws.title}] {oc.coordinate}: {mask_of(ov)!r} -> {mask_of(rv)!r}")
                if ov == rv:
                    all_unchanged.append(f"[{o_ws.title}] {oc.coordinate}: {ov!r}")
                else:
                    changed += 1
                if entry.get("type") == "name" and name_template(ov) != name_template(rv):
                    all_tmpl.append(f"[{o_ws.title}] {oc.coordinate}: {ov!r} -> {rv!r}")
                if len(ov.strip()) >= 3 and entry.get("type") not in _SMALL_SPACE_TYPES:
                    originals.add(ov.strip())
                if oc.number_format != rc.number_format:
                    all_mask.append(f"[{o_ws.title}] {oc.coordinate}: number format changed")
        # Scan the whole used range, not just the data rows. Title and banner rows
        # above the header are a favourite place for "Patient List -- A. Test", and
        # they are copied through verbatim by a redactor that starts below the
        # header.
        r_texts.append("\n".join("\x1f".join("" if c.value is None else str(c.value) for c in row)
                                for row in r_ws.iter_rows(min_row=1)))

    rep.add("no PHI-shaped column is missing from the plan", not suspects, _sample(suspects, 4))
    unmatched = sorted(
        e["name"] for e in plan.get("columns", [])
        if not ({e["name"].strip().lower()} | {a.strip().lower() for a in e.get("aliases", [])})
        & seen_headers
    )
    rep.add("every plan column matched a header in the workbook", not unmatched,
            f"unmatched: {_sample(unmatched, 5)}" if unmatched else f"{len(seen_headers)} headers seen")

    rep.add("non-PHI cells unchanged", not all_nonphi, _sample(all_nonphi))
    rep.add("PHI cell masks/types/formats preserved", not all_mask, _sample(all_mask))
    rep.add("no PHI cell left unchanged", not all_unchanged, _sample(all_unchanged))
    rep.add("name templates preserved", not all_tmpl, _sample(all_tmpl))
    rep.add("PHI cells were rewritten", changed > 0 or total == 0, f"{changed}/{total} changed")

    haystack = "\n".join(r_texts)
    leaked = sorted(v for v in originals if v in haystack)
    rep.add("no original PHI value appears anywhere in output", not leaked,
            f"{len(leaked)} leaked: {_sample(leaked, 5)}" if leaked else
            f"checked {len(originals)} distinct original values")


def report_distributions(orig: Path, red: Path, plan: dict) -> str:
    """Show template frequencies side by side, for the human to eyeball.

    Preserving each row's own template should make these match exactly. Printing
    them is how that claim gets checked rather than assumed.
    """
    if orig.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        return "(distribution report supported for delimited files)"
    o_rows, _ = _read_delimited(orig, plan)
    r_rows, _ = _read_delimited(red, plan)
    hr = plan.get("header_row", 0)
    header = [h.strip().lower() for h in o_rows[hr]]
    phi_cols, _named = _plan_cols(plan)
    out = []
    for j, col in enumerate(header):
        entry = phi_cols.get(col)
        if not entry or entry.get("type") not in {"name", "date", "id", "ssn", "phone"}:
            continue
        kind = entry["type"]
        shaper = name_template if kind == "name" else mask_of
        oc = Counter(shaper(r[j].strip()) for r in o_rows[hr + 1:] if j < len(r) and r[j].strip())
        rc = Counter(shaper(r[j].strip()) for r in r_rows[hr + 1:] if j < len(r) and r[j].strip())
        total = sum(oc.values()) or 1
        out.append(f"\n{o_rows[hr][j]}  ({kind}) -- {'MATCH' if oc == rc else 'DIFFERS'}")
        for tmpl, n in oc.most_common(8):
            out.append(f"    {tmpl:<28} orig {n:>4} ({100 * n / total:5.1f}%)   redacted {rc.get(tmpl, 0):>4}")
    return "\n".join(out) if out else "(no distribution-tracked columns in plan)"


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--original", required=True, type=Path)
    ap.add_argument("--redacted", required=True, type=Path)
    ap.add_argument("--plan", type=Path)
    ap.add_argument("--distributions", action="store_true", help="print template frequency comparison")
    ap.add_argument("--json", action="store_true")
    args = ap.parse_args()

    plan = json.loads(args.plan.read_text()) if args.plan else {"columns": []}
    rep = Report()

    rep.add("extension identical", args.original.suffix.lower() == args.redacted.suffix.lower(),
            f"{args.original.suffix} vs {args.redacted.suffix}")
    rep.add("filename recognisably related",
            args.original.stem.split("_")[0][:6].lower() in args.redacted.stem.lower(),
            f"{args.original.name} -> {args.redacted.name}", advisory=True)

    if args.original.suffix.lower() == ".xls":
        # Mirrors redact.py: openpyxl cannot read legacy BIFF, and its failure deep
        # in the zip reader reads like a corrupt file rather than an unsupported one.
        print("error: legacy .xls is not supported -- re-save as .xlsx first", file=sys.stderr)
        return 2
    if args.original.suffix.lower() in {".xlsx", ".xlsm"}:
        verify_excel(args.original, args.redacted, plan, rep)
    else:
        verify_delimited(args.original, args.redacted, plan, rep)

    if args.json:
        print(json.dumps({"ok": rep.ok, "checks": rep.checks}, indent=2))
    else:
        print(rep.render())
        if args.distributions:
            print("\n--- template distributions ---")
            print(report_distributions(args.original, args.redacted, plan))
    return 0 if rep.ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
